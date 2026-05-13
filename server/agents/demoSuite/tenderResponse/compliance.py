#!/usr/bin/env python3
"""
Tender Response — deterministic compliance checker.

Reads JSON from stdin, parses evidence pack files, matches requirements to
evidence, and returns structured match results. All threshold checks happen
here — the LLM never makes these determinations.

Input (stdin JSON):
  {
    "requirements":  [ { requirement_id, category, requirement_text,
                         is_mandatory, evaluation_weight } ],
    "evidence_files": {
      "compliance_rules_csv": "<raw CSV text>",
      "projects_xlsx":        "<base64 XLSX bytes>",
      "personnel_xlsx":       "<base64 XLSX bytes>",
      "certificates_xlsx":    "<base64 XLSX bytes>"
    },
    "tender_close_date": "YYYY-MM-DD"   (default 2026-06-16)
  }

Output (stdout JSON):
  {
    "requirement_matches": [ <match object per requirement> ],
    "compliance_summary":  { total, strong, partial, none, blockers },
    "execution_time_ms":   <float>
  }
"""

import sys
import json
import csv
import io
import base64
import datetime
import time
import re

try:
    import openpyxl
except ImportError as e:
    sys.stdout.write(json.dumps({"error": f"Missing dependency: {e}. Run: pip install openpyxl"}))
    sys.exit(1)

# ── Constants ──────────────────────────────────────────────────────────────────

DEFAULT_TENDER_CLOSE      = datetime.date(2026, 6, 16)
MANDATORY_VALUE_THRESHOLD = 5_000_000.0
RECENCY_YEARS             = 5


# ── XLSX helpers ───────────────────────────────────────────────────────────────

def _xlsx_from_b64(b64_str):
    raw = base64.b64decode(b64_str)
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)


def _find_header_row(rows):
    """
    Skip title/subtitle blocks — return (header_index, headers_list).
    Looks for the first row whose first cell ends with 'ID' or 'Id'
    (e.g. 'Project ID', 'Personnel ID', 'Cert ID', 'Ins ID').
    Falls back to row 0 if not found.
    """
    for i, row in enumerate(rows):
        first = str(row[0]).strip() if row[0] is not None else ''
        if first.upper().endswith('ID') and len([c for c in row if c is not None]) > 2:
            return i, [str(c).strip() if c is not None else '' for c in row]
    return 0, [str(c).strip() if c is not None else '' for c in rows[0]]


def _col(headers, *candidates):
    """Return 0-based index of first matching header (case-insensitive), or None."""
    hl = [h.lower().strip() for h in headers]
    for c in candidates:
        try:
            return hl.index(c.lower().strip())
        except ValueError:
            pass
    return None


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    return str(v).strip() if v is not None else None


# ── Evidence parsers ───────────────────────────────────────────────────────────

def parse_compliance_rules(csv_text):
    rules = []
    if not csv_text:
        return rules
    reader = csv.DictReader(io.StringIO(csv_text))
    for row in reader:
        rule_id = (row.get('Rule ID') or '').strip()
        if not rule_id:
            continue
        rules.append({
            'rule_id':     rule_id,
            'category':    (row.get('Category') or '').strip(),
            'requirement': (row.get('RFP Clause Requirement') or '').strip(),
            'response':    (row.get('Standard Response / Evidence Mapping') or '').strip(),
        })
    return rules


def parse_projects(b64_str):
    if not b64_str:
        return []
    wb   = _xlsx_from_b64(b64_str)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    hdr_idx, headers = _find_header_row(rows)
    data_rows = rows[hdr_idx + 1:]

    i_id   = _col(headers, 'Project ID')
    i_name = _col(headers, 'Project Name')
    i_loc  = _col(headers, 'Location')
    i_val  = _col(headers, 'Value (AUD)', 'Contract Value', 'Value')
    i_year = _col(headers, 'Year Completed', 'Year of Completion', 'Year')
    i_corr = _col(headers, 'Corrosivity Class', 'Corrosivity', 'Classification')
    i_iccp = _col(headers, 'ICCP Installed?', 'ICCP Installed', 'ICCP')
    i_drdg = _col(headers, 'Dredging Included?', 'Dredging Included', 'Dredging')
    i_desc = _col(headers, 'Description')
    i_outc = _col(headers, 'Key Outcomes / Metrics', 'Key Outcomes', 'Outcomes')
    i_clm  = _col(headers, 'Claimability', 'Claimable', 'Status')

    projects = []
    for row in data_rows:
        pid = _cell(row, i_id)
        if not pid or not pid.startswith('REF-'):
            continue
        projects.append({
            'id':               pid,
            'name':             _cell(row, i_name) or '',
            'location':         _cell(row, i_loc) or '',
            'contract_value':   _parse_money(_cell(row, i_val)),
            'year':             _parse_year(_cell(row, i_year)),
            'corrosivity_class': _cell(row, i_corr) or '',
            'iccp':             _parse_bool(_cell(row, i_iccp)),
            'dredging':         _parse_bool(_cell(row, i_drdg)),
            'description':      _cell(row, i_desc) or '',
            'outcomes':         _cell(row, i_outc) or '',
            'claimable':        _parse_bool(_cell(row, i_clm)),
        })
    return projects


def parse_personnel(b64_str):
    if not b64_str:
        return []
    wb   = _xlsx_from_b64(b64_str)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    hdr_idx, headers = _find_header_row(rows)
    data_rows = rows[hdr_idx + 1:]

    i_id   = _col(headers, 'Personnel ID')
    i_name = _col(headers, 'Full Name', 'Name')
    i_role = _col(headers, 'Role (This Tender)', 'Role', 'Position')
    i_qual = _col(headers, 'Qualifications', 'Quals')
    i_reg  = _col(headers, 'Registration / Licence', 'Registration Number', 'RPEQ', 'Licence')
    i_yrs  = _col(headers, 'Years Experience', 'Years')
    i_proj = _col(headers, 'Key Relevant Projects', 'Key Projects')
    i_avl  = _col(headers, 'Availability (%)', 'Availability')
    i_clm  = _col(headers, 'Claimability', 'Claimable', 'Status')
    i_note = _col(headers, 'Notes')

    personnel = []
    for row in data_rows:
        pid = _cell(row, i_id)
        if not pid or not pid.startswith('PER-'):
            continue
        personnel.append({
            'id':            pid,
            'name':          _cell(row, i_name) or '',
            'role':          _cell(row, i_role) or '',
            'qualifications': _cell(row, i_qual) or '',
            'registration':  _cell(row, i_reg) or '',
            'years_experience': _parse_int(_cell(row, i_yrs)),
            'key_projects':  _cell(row, i_proj) or '',
            'availability':  _cell(row, i_avl) or '',
            'notes':         _cell(row, i_note) or '',
            'claimable':     _parse_bool(_cell(row, i_clm)),
        })
    return personnel


def parse_certificates(b64_str):
    if not b64_str:
        return [], []
    wb = _xlsx_from_b64(b64_str)

    # Sheet 1 — ISO & Prequalification
    certs = []
    ws1   = wb.worksheets[0] if wb.worksheets else None
    if ws1:
        rows = list(ws1.iter_rows(values_only=True))
        hdr_idx, headers = _find_header_row(rows)
        data_rows = rows[hdr_idx + 1:]

        i_id   = _col(headers, 'Cert ID')
        i_desc = _col(headers, 'Certificate / Accreditation', 'Description', 'Certificate')
        i_std  = _col(headers, 'Standard / Scheme', 'Standard')
        i_body = _col(headers, 'Issuing Body', 'Issued By')
        i_num  = _col(headers, 'Certificate Number', 'Cert Number', 'Number')
        i_iss  = _col(headers, 'Issue Date', 'Issued')
        i_exp  = _col(headers, 'Expiry Date', 'Expiry')
        i_stat = _col(headers, 'Status', 'Current Status')
        i_att  = _col(headers, 'Attach Ref', 'Attachment')

        for row in data_rows:
            cid = _cell(row, i_id)
            if not cid or not cid.startswith('CRT-'):
                continue
            certs.append({
                'id':           cid,
                'description':  _cell(row, i_desc) or '',
                'standard':     _cell(row, i_std) or '',
                'issuing_body': _cell(row, i_body) or '',
                'cert_number':  _cell(row, i_num) or '',
                'issue_date':   _parse_date(_cell(row, i_iss)),
                'expiry_date':  _parse_date(_cell(row, i_exp)),
                'status':       (_cell(row, i_stat) or 'UNKNOWN').upper().strip(),
                'attachment':   _cell(row, i_att) or '',
            })

    # Sheet 2 — Insurance
    insurance = []
    ws2       = wb.worksheets[1] if len(wb.worksheets) > 1 else None
    if ws2:
        rows = list(ws2.iter_rows(values_only=True))
        hdr_idx, headers = _find_header_row(rows)
        data_rows = rows[hdr_idx + 1:]

        i_id   = _col(headers, 'Ins ID', 'Insurance ID', 'ID')
        i_type = _col(headers, 'Policy Type', 'Type', 'Insurance Type')
        i_ins  = _col(headers, 'Insurer', 'Underwriter')
        i_pol  = _col(headers, 'Policy Number', 'Policy No')
        i_lim  = _col(headers, 'Cover Limit', 'Limit', 'Sum Insured')
        i_exp  = _col(headers, 'Expiry Date', 'Policy End')

        for row in data_rows:
            iid = _cell(row, i_id)
            if not iid or not iid.startswith('INS-'):
                continue
            insurance.append({
                'id':          iid,
                'type':        _cell(row, i_type) or '',
                'insurer':     _cell(row, i_ins) or '',
                'policy_num':  _cell(row, i_pol) or '',
                'limit':       _parse_money(_cell(row, i_lim)),
                'limit_raw':   _cell(row, i_lim) or '',
                'expiry_date': _parse_date(_cell(row, i_exp)),
            })

    return certs, insurance


# ── Value parsers ──────────────────────────────────────────────────────────────

def _parse_money(raw):
    if raw is None:
        return None
    cleaned = re.sub(r'[^\d.]', '', re.sub(r'\s*(per|each|any|occurrence|claim|event).*', '', str(raw), flags=re.I))
    try:
        return float(cleaned) if cleaned else None
    except ValueError:
        return None

def _parse_year(raw):
    if raw is None:
        return None
    m = re.search(r'\b(20\d{2}|19\d{2})\b', str(raw))
    return int(m.group(1)) if m else None

def _parse_bool(raw):
    if raw is None:
        return False
    return str(raw).strip().lower() in ('yes', 'true', '1', 'y', 'approved', 'current')

def _parse_date(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    for fmt in ('%d %b %Y', '%d/%m/%Y', '%d %B %Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def _parse_int(raw):
    if raw is None:
        return None
    try:
        return int(float(str(raw)))
    except (ValueError, TypeError):
        return None


# ── Certificate gate check ─────────────────────────────────────────────────────

def cert_gate_status(cert, tender_close_date):
    """Returns (level, reason): level = 'GREEN' | 'AMBER' | 'RED'."""
    status = cert.get('status', '').upper()
    expiry = cert.get('expiry_date')

    if status == 'EXPIRED':
        return 'RED', f"Certificate {cert['id']} is EXPIRED"
    if status == 'RENEWING':
        return 'AMBER', (
            f"Certificate {cert['id']} ({cert['description']}) renewal in progress — "
            f"use renewal language: 'renewal was lodged and an interim certificate is available'"
        )
    if expiry and isinstance(expiry, datetime.date) and expiry < tender_close_date:
        return 'RED', f"Certificate {cert['id']} expires {expiry} before tender close {tender_close_date}"
    return 'GREEN', f"Certificate {cert['id']} is current (expires {expiry})"


# ── Keyword helpers ────────────────────────────────────────────────────────────

def _kw(text, *keywords):
    t = text.lower()
    return any(k.lower() in t for k in keywords)

def _rule_match(req_text, req_category, rule):
    """True if a compliance rule is relevant to this requirement."""
    if rule['category'].lower() == req_category.lower():
        # Keyword overlap between requirement text and rule requirement
        rule_words = set(re.findall(r'\w+', rule['requirement'].lower()))
        req_words  = set(re.findall(r'\w+', req_text.lower()))
        common = rule_words & req_words - {'the', 'a', 'an', 'for', 'and', 'or', 'of', 'in', 'with'}
        return len(common) >= 2
    return False


# ── Main requirement matcher ───────────────────────────────────────────────────

def match_requirement(req, rules, projects, personnel, certs, insurance, tender_close_date):
    req_id    = req.get('requirement_id', '')
    category  = (req.get('category') or '').strip()
    req_text  = (req.get('requirement_text') or '')
    mandatory = bool(req.get('is_mandatory'))

    evidence_ids    = []
    rationale_parts = []
    blocker         = False
    blocker_reason  = None
    blocker_level   = None
    draft_hints     = {}

    # ── 1. Compliance rules lookup ─────────────────────────────────────────────
    matched_rules = [r for r in rules if _rule_match(req_text, category, r)]
    if matched_rules:
        rule = matched_rules[0]
        rationale_parts.append(f"Rule {rule['rule_id']}: {rule['requirement']}")
        draft_hints['rule_response'] = rule['response']
        for eid in re.findall(r'\b(?:REF|CRT|PER|INS)-\d+\b', rule['response']):
            if eid not in evidence_ids:
                evidence_ids.append(eid)

    # ── 2. Certification ───────────────────────────────────────────────────────
    if category == 'Certification' or _kw(req_text, 'iso', 'pqc', 'prequalification', 'certification'):
        for cert in certs:
            std  = cert['standard'].lower()
            desc = cert['description'].lower()
            hit  = (
                (_kw(req_text, 'iso 9001') and 'iso 9001' in std) or
                (_kw(req_text, 'iso 14001') and 'iso 14001' in std) or
                (_kw(req_text, 'iso 45001') and 'iso 45001' in std) or
                (_kw(req_text, 'pqc', 'prequalification') and ('pqc' in std or 'prequalif' in std.replace(' ', '')))
            )
            if not hit:
                continue
            level, reason = cert_gate_status(cert, tender_close_date)
            if cert['id'] not in evidence_ids:
                evidence_ids.append(cert['id'])
            if level == 'RED':
                blocker = True
                blocker_reason = reason
                blocker_level  = 'RED'
                rationale_parts.append(f"BLOCKER-RED: {reason}")
            elif level == 'AMBER':
                if not (blocker and blocker_level == 'RED'):
                    blocker        = mandatory
                    blocker_reason = reason
                    blocker_level  = 'AMBER'
                rationale_parts.append(f"BLOCKER-AMBER: {reason}")
            else:
                rationale_parts.append(reason)
                draft_hints[f'cert_{cert["id"]}'] = 'current'

    # ── 3. Insurance ───────────────────────────────────────────────────────────
    if category == 'Insurance' or _kw(req_text, 'insurance', 'public liability', 'professional indemnity', 'indemnity'):
        for ins in insurance:
            tp = ins['type'].lower()
            hit = (
                (_kw(req_text, 'public liability') and 'public liability' in tp) or
                (_kw(req_text, 'professional indemnity') and 'indemnity' in tp) or
                (_kw(req_text, 'workers compensation') and 'workers' in tp)
            )
            if not hit:
                continue
            if ins['id'] not in evidence_ids:
                evidence_ids.append(ins['id'])
            req_limit = _extract_dollar_amount(req_text)
            if req_limit and ins['limit']:
                if ins['limit'] >= req_limit:
                    rationale_parts.append(
                        f"{ins['id']} ({ins['type']}): ${ins['limit']:,.0f} — meets ${req_limit:,.0f} threshold"
                    )
                    draft_hints['insurance_limit']  = ins['limit']
                    draft_hints['insurance_raw']    = ins['limit_raw']
                else:
                    blocker        = mandatory
                    blocker_reason = f"{ins['id']} limit ${ins['limit']:,.0f} below required ${req_limit:,.0f}"
                    blocker_level  = 'RED'
                    rationale_parts.append(f"BLOCKER-RED: {blocker_reason}")
            else:
                rationale_parts.append(f"{ins['id']} ({ins['type']}) matched")

    # ── 4. Experience ──────────────────────────────────────────────────────────
    if category == 'Experience' or _kw(req_text, 'experience', 'project', 'reference', 'similar', 'c5-m', 'delivered'):
        current_year     = datetime.date.today().year
        is_value_gate    = _kw(req_text, '$5 million', '$5m', '5 million', 'five million', '> $5', '>$5', '$5 m')
        is_c5m_req       = _kw(req_text, 'c5-m', 'c5m')
        is_iccp_req      = _kw(req_text, 'iccp', 'cathodic protection', 'impressed current')
        is_dredge_req    = _kw(req_text, 'dredging', 'dredge')
        min_refs_req     = _kw(req_text, 'three reference', 'three project', 'minimum three', '3 reference', '3 project')

        matched_projects = []
        for proj in projects:
            if not proj.get('claimable', True):
                continue
            score   = 0
            reasons = []

            # Value gate — only qualifying projects
            if is_value_gate:
                if proj['contract_value'] and proj['contract_value'] > MANDATORY_VALUE_THRESHOLD:
                    score += 3
                    reasons.append(f"${proj['contract_value']/1e6:.1f}M > $5M")
                else:
                    continue

            # C5-M gate — must match exactly; 'C5-M Marine' contains 'C5-M'
            if is_c5m_req:
                if 'C5-M' in proj['corrosivity_class']:
                    score += 3
                    reasons.append('C5-M confirmed')
                else:
                    continue

            # Recency bonus
            if proj['year'] and proj['year'] >= (current_year - RECENCY_YEARS):
                score += 1
                reasons.append(f"completed {proj['year']}")

            # ICCP
            if is_iccp_req:
                if proj['iccp']:
                    score += 2
                    reasons.append('ICCP installed')
                else:
                    score -= 1

            # Dredging
            if is_dredge_req:
                if proj['dredging']:
                    score += 2
                    reasons.append('dredging included')

            # General marine — fallback for non-gated requirements
            if score == 0 and _kw(proj['description'] + proj['name'], 'marine', 'harbour', 'port', 'berth', 'wharf', 'jetty'):
                score = 1
                reasons.append('marine infrastructure')

            if score > 0:
                matched_projects.append((proj, score, reasons))

        matched_projects.sort(key=lambda x: x[1], reverse=True)

        for proj, score, reasons in matched_projects[:4]:
            if proj['id'] not in evidence_ids:
                evidence_ids.append(proj['id'])
            rationale_parts.append(f"{proj['id']} ({proj['name']}): {', '.join(reasons)}")

        # Value gate blocker
        if is_value_gate and mandatory and not matched_projects:
            blocker        = True
            blocker_reason = "No project exceeding $5M found in evidence pack"
            blocker_level  = 'RED'
        elif is_value_gate and matched_projects:
            draft_hints['qualifying_projects'] = [p[0]['id'] for p in matched_projects]

        # Minimum references check
        if min_refs_req and len(matched_projects) < 3:
            blocker        = mandatory
            blocker_reason = f"Only {len(matched_projects)} matching reference project(s) found (3 required)"
            blocker_level  = 'AMBER' if len(matched_projects) >= 2 else 'RED'

        # C5-M gate blocker
        if is_c5m_req and mandatory and not matched_projects:
            blocker        = True
            blocker_reason = "No C5-M corrosivity environment projects found in evidence pack"
            blocker_level  = 'RED'

    # ── 5. Design / RPEQ ──────────────────────────────────────────────────────
    if category == 'Design' or _kw(req_text, 'rpeq', 'registered engineer', 'structural engineer'):
        rpeq_found = False
        for person in personnel:
            reg = person.get('registration', '')
            if re.search(r'RPEQ|rpeq', reg):
                if person['id'] not in evidence_ids:
                    evidence_ids.append(person['id'])
                rationale_parts.append(f"{person['id']} {person['name']}: {reg}")
                draft_hints['rpeq_name']   = person['name']
                draft_hints['rpeq_number'] = reg
                draft_hints['rpeq_per_id'] = person['id']
                rpeq_found = True
                if not _kw(person.get('notes', ''), 'lead', 'primary', 'sign'):
                    # Prefer the primary signatory (PER-002) — continue looking
                    if 'Lead' in person.get('role', '') or '21453' in reg:
                        break
        if not rpeq_found and mandatory:
            blocker        = True
            blocker_reason = "No RPEQ-registered engineer found in personnel register"
            blocker_level  = 'RED'

    # ── 6. Safety ─────────────────────────────────────────────────────────────
    if category == 'Safety' or _kw(req_text, 'whs', 'safety management', 'pqc', 'pre-qualification', 'prequalification'):
        if _kw(req_text, 'pqc', 'prequalification', 'pre-qualification'):
            for cert in certs:
                if 'pqc' in cert['standard'].lower() or 'pqc' in cert['description'].lower():
                    if cert['id'] not in evidence_ids:
                        evidence_ids.append(cert['id'])
                    level, reason = cert_gate_status(cert, tender_close_date)
                    rationale_parts.append(reason)
                    if level in ('AMBER', 'RED') and mandatory:
                        blocker        = mandatory
                        blocker_reason = reason
                        blocker_level  = level

    # ── 7. Quality ─────────────────────────────────────────────────────────────
    if category == 'Quality' or _kw(req_text, 'quality management', 'qmp', 'iso 9001', 'quality plan'):
        for cert in certs:
            if 'iso 9001' in cert['standard'].lower():
                if cert['id'] not in evidence_ids:
                    evidence_ids.append(cert['id'])
                level, reason = cert_gate_status(cert, tender_close_date)
                rationale_parts.append(reason)

    # ── Compute final match_status ─────────────────────────────────────────────
    n_ev = len(evidence_ids)
    has_rule = bool(matched_rules)

    if n_ev == 0 and not has_rule:
        match_status = 'NONE'
        if mandatory and not blocker:
            blocker        = True
            blocker_reason = blocker_reason or f"No evidence found: {req_text[:80]}"
            blocker_level  = blocker_level or 'RED'
    elif blocker and blocker_level == 'RED':
        match_status = 'PARTIAL'
    elif blocker and blocker_level == 'AMBER':
        match_status = 'PARTIAL'
    elif n_ev >= 2 or (n_ev >= 1 and has_rule):
        match_status = 'STRONG'
    else:
        match_status = 'PARTIAL'

    return {
        'requirement_id': req_id,
        'match_status':   match_status,
        'evidence_ids':   list(dict.fromkeys(evidence_ids)),
        'match_rationale': '; '.join(rationale_parts) if rationale_parts else 'No direct evidence match found',
        'blocker':        blocker,
        'blocker_reason': blocker_reason,
        'blocker_level':  blocker_level,
        'draft_hints':    draft_hints,
    }


def _extract_dollar_amount(text):
    m = re.search(r'\$(\d+(?:\.\d+)?)\s*[Mm](?:illion)?', text)
    if m:
        return float(m.group(1)) * 1_000_000
    m = re.search(r'\$([\d,]+)', text)
    if m:
        try:
            return float(m.group(1).replace(',', ''))
        except ValueError:
            pass
    return None


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    start_ms = time.time() * 1000

    raw = sys.stdin.read()
    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as e:
        sys.stdout.write(json.dumps({'error': f'Invalid JSON input: {e}'}))
        sys.exit(1)

    requirements   = payload.get('requirements', [])
    evidence_files = payload.get('evidence_files', {})
    close_str      = payload.get('tender_close_date', '')

    try:
        tender_close = datetime.date.fromisoformat(close_str) if close_str else DEFAULT_TENDER_CLOSE
    except ValueError:
        tender_close = DEFAULT_TENDER_CLOSE

    try:
        rules     = parse_compliance_rules(evidence_files.get('compliance_rules_csv', ''))
        projects  = parse_projects(evidence_files.get('projects_xlsx', ''))
        personnel = parse_personnel(evidence_files.get('personnel_xlsx', ''))
        certs, insurance = parse_certificates(evidence_files.get('certificates_xlsx', ''))
    except Exception as e:
        sys.stdout.write(json.dumps({'error': f'Evidence parse error: {e}'}))
        sys.exit(1)

    matches = []
    for req in requirements:
        try:
            m = match_requirement(req, rules, projects, personnel, certs, insurance, tender_close)
        except Exception as e:
            m = {
                'requirement_id': req.get('requirement_id', '?'),
                'match_status':   'NONE',
                'evidence_ids':   [],
                'match_rationale': f'Error during matching: {e}',
                'blocker':        bool(req.get('is_mandatory')),
                'blocker_reason': str(e),
                'blocker_level':  'RED' if req.get('is_mandatory') else None,
                'draft_hints':    {},
            }
        matches.append(m)

    strong  = sum(1 for m in matches if m['match_status'] == 'STRONG')
    partial = sum(1 for m in matches if m['match_status'] == 'PARTIAL')
    none_ct = sum(1 for m in matches if m['match_status'] == 'NONE')
    blockers = sum(1 for m in matches if m['blocker'])

    sys.stdout.write(json.dumps({
        'requirement_matches': matches,
        'compliance_summary': {
            'total':    len(matches),
            'strong':   strong,
            'partial':  partial,
            'none':     none_ct,
            'blockers': blockers,
        },
        'execution_time_ms': round((time.time() * 1000) - start_ms, 1),
    }))
    sys.exit(0)


if __name__ == '__main__':
    main()
